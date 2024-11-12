import random
from data import data_dictionary
from openpyxl import Workbook,load_workbook

def generate_npc(data_dictonary):


      weights = data_dictonary["tribe_weights"]
      total_weight = sum(weights)
      weights = [weight / total_weight * 100 for weight in weights]

      npc_dictionary = {
      "npc" : random.choice(data_dictonary["npc"]),
      "tribe" : random.choices(data_dictonary["tribe"], weights=weights, k=1)[0],
      "personality" : random.choice(data_dictonary["personality"]),
      "voice" : random.choice(data_dictonary["voice"]),
      "speech" : random.choice(data_dictonary["speech"]),
      "catchphrase" : random.choice(data_dictonary["catchphrase"]),
      "quirks" : random.choice(data_dictonary["quirks"]),
      "appearance" : random.choice(data_dictonary["appearance"]),
      "statement_piece" : random.choice(data_dictonary["statement_piece"]),
      "interests" : random.choice(data_dictonary["interests"]),
      "fighting_style" : random.choice(data_dictonary["fighting_style"]),
      }

      for key,value in npc_dictionary.items():
            print(f"{key}: {value}")

def export_to_xlsx(data_dictonary):

      # Initialize a new workbook
      workbook = Workbook()
      # Remove the default sheet created by openpyxl
      workbook.remove(workbook.active)

      # Populate each key's data in a separate sheet
      for key, values in data_dictionary.items():
            # Create a new sheet with the name of the key
            sheet = workbook.create_sheet(title=key)
      
            # Write each item in the list to a new row in the sheet
            for row_num, value in enumerate(values, start=1):
                  sheet.cell(row=row_num, column=1, value=value)

      # Save the workbook to a file
      workbook.save("data_dictionary.xlsx")

def import_from_xlsx(file_name):
      # Load the workbook
      workbook = load_workbook(file_name)

      # Initialize a dictionary to hold the data
      data_dictionary = {}

      # Iterate through each sheet in the workbook
      for sheet_name in workbook.sheetnames:
            # Get the current sheet
            sheet = workbook[sheet_name]
      
            # Extract data from each row in column A and add to list
            data_dictionary[sheet_name] = [cell.value for cell in sheet['A'] if cell.value is not None]

      # Close the workbook after extraction
      workbook.close()

      # Your data is now in data_dictionary
      return data_dictionary
